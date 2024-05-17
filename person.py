import openpyxl
import datetime
import calendar
import os


def _parse_date(date_str):
    if not date_str:
        return None

    formats = ["%d.%m.%Y", "%d %m %Y", "%d/%m/%Y", "%d,%m,%Y", "%d-%m-%Y"]

    for date_format in formats:
        try:
            return datetime.datetime.strptime(date_str, date_format).date()
        except ValueError:
            continue

    raise ValueError("Invalid date format. Please use one of the following formats: %s" % ", ".join(formats))


def count_leap_years(start_year, end_year):
    leap_years = sum(1 for year in range(start_year, end_year + 1) if calendar.isleap(year))
    return leap_years


class Person:
    def __init__(self, first_name, last_name, birth_date, gender, middle_name=None, death_date=None):
        self.first_name = first_name
        self.last_name = last_name
        self.middle_name = middle_name
        self.birth_date = _parse_date(birth_date)
        self.death_date = _parse_date(death_date) if death_date else None
        self.gender = gender

    def age(self, today=None):
        if not today:
            today = datetime.date.today()

        age = today.year - self.birth_date.year

        if (self.birth_date.month, self.birth_date.day) > (today.month, today.day):
            age -= 1
        return age

    def to_excel_row(self):
        return [self.first_name, self.last_name, self.middle_name, self.birth_date.strftime('%d.%m.%Y'),
                self.death_date.strftime('%d.%m.%Y') if self.death_date else '', self.gender]

class PersonDatabase:
    def __init__(self):
        self.people = []

    def add_person(self, person):
        self.people.append(person)

    def delete_person(self, person):
        self.people.remove(person)

    def search_person(self, query):
        results = []
        for person in self.people:
            if (query.lower() in person.first_name.lower() or
                    query.lower() in person.last_name.lower() or
                    (person.middle_name and query.lower() in person.middle_name.lower()) or
                    query.lower() == person.gender.lower()):
                results.append(person)
            elif any(part.lower() in person.first_name.lower() for part in query.split()) or \
                    any(part.lower() in person.last_name.lower() for part in query.split()) or \
                    (person.middle_name and any(part.lower() in person.middle_name.lower() for part in query.split())):
                results.append(person)
        return results

    def save_to_excel(self, filename):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["First Name", "Last Name", "Middle Name", "Birth Date", "Death Date", "Gender"])
        for person in self.people:
            sheet.append(person.to_excel_row())
        workbook.save(filename)

    def load_from_excel(self, filename):
        while True:
            try:
                workbook = openpyxl.load_workbook(filename)
                sheet = workbook.active
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    first_name, last_name, middle_name, birth_date, death_date, gender = row
                    self.add_person(Person(first_name, last_name, birth_date, gender, middle_name, death_date))
                break
            except FileNotFoundError:
                print("File not found.")
                filename = input("Enter filename to load: ")
                full_path = os.path.join("D:\\", filename + ".xlsx")
                continue
            except Exception as e:
                print("An error occurred while loading the database:", e)
                break
